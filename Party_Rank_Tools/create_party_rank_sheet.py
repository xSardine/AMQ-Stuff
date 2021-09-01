import json
import re
from typing import List
import PR_config
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.styles.colors import Color
from openpyxl.styles.alignment import Alignment

# todo arial not working

party_rank_name = PR_config.party_rank_name

# Filtering search
anime_search_filters = PR_config.anime_search_filters
artist_search_filters = PR_config.artist_search_filters
song_name_search_filters = PR_config.song_name_search_filters
# Filtering search

# Sheet Configuration
file_name = party_rank_name + " Anime Songs Ranking Sheet.xlsx"
sheet_name = "Sheet1"
link_color = "1155cc"
cell_background_color = "cccccc"
border_color = "949494"
font_police = "Arial"
first_line_font_size = 12
rest_font_size = 11
# Sheet Configuration

and_filter = PR_config.and_filter
filter_duplicate = PR_config.filter_duplicate

# End of configuration

ANIME_REGEX_REPLACE_RULES = [
    {"input": "ou", "replace": "(ou|ō)"},
    {"input": "oo", "replace": "(oo|ō)"},
    {"input": "o", "replace": "[oōóòöôøΦ]"},
    {"input": "u", "replace": "([uūûúùüǖ]|uu)"},
    {"input": "a", "replace": "[aä@âàáạåæā]"},
    {"input": "c", "replace": "[cč]"},
    {"input": " ", "replace": "( ?[★☆\\/\\*=\\+·♥∽・〜†×♪→␣:;~\\-?,.!@_]+ ?| )"},
    {"input": "e", "replace": "[eéêëèæē]"},
    {"input": "'", "replace": "['’]"},
    {"input": "n", "replace": "[nñ]"},
    {"input": "2", "replace": "[2²]"},
    {"input": "i", "replace": "[ií]"},
    {"input": "3", "replace": "[3³]"},
    {"input": "x", "replace": "[x×]"},
    {"input": "b", "replace": "[bß]"},
]


def escapeRegExp(str):
    str = re.escape(str)
    str = str.replace("\ ", " ")
    return str


def replace_regex_values(search_filters):

    filters_regexs = []
    for filter in search_filters:
        exact_flag = True if isinstance(filter, list) and filter[1] else False
        filter = filter[0] if isinstance(filter, list) else filter
        filter = escapeRegExp(filter)
        for rule in ANIME_REGEX_REPLACE_RULES:
            filter = filter.replace(rule["input"], rule["replace"])
            if exact_flag:
                filters_regexs.append("^" + filter + "$")
            else:
                filters_regexs.append(".*" + filter + ".*")
    return filters_regexs


def anime_contains_anime_filters(anime, anime_search_filters):

    if len(anime_search_filters) == 0:
        return True if and_filter else False
    else:
        for filter in anime_search_filters:
            if re.match(filter, anime["name"], re.IGNORECASE):
                return True
    return False


def anime_contains_song_filters(song, artist_search_filters, song_name_search_filters):

    flag_artist = False
    flag_song_name = False

    if len(artist_search_filters) > 0:
        for filter in artist_search_filters:
            if re.match(filter, song["artist"], re.IGNORECASE):
                flag_artist = True
    else:
        flag_artist = True if and_filter else False

    if len(song_name_search_filters) > 0:
        for filter in song_name_search_filters:
            if re.match(filter, song["name"], re.IGNORECASE):
                flag_song_name = True
    else:
        flag_song_name = True if and_filter else False

    return (
        flag_artist and flag_song_name if and_filter else flag_artist or flag_song_name
    )


def format_song(song):

    if song["type"] == 1:
        type = "OP"
    elif song["type"] == 2:
        type = "ED"
    else:
        type = "IN"
    number = song["number"] if song["number"] != 0 else ""

    link = (
        song["examples"]["720"]
        if "720" in song["examples"].keys()
        else song["examples"]["480"]
        if "480" in song["examples"].keys()
        else song["examples"]["mp3"]
        if "mp3" in song["examples"].keys()
        else "Not Uploaded"
    )

    mp3_link = (
        song["examples"]["mp3"]
        if "mp3" in song["examples"].keys()
        else "https://www.youtube.com/watch?v=dQw4w9WgXcQ"
    )

    return {
        "type": type + str(number),
        "info": '"' + song["name"] + '"' + " by " + song["artist"],
        "link": link,
        "mp3_link": mp3_link,
    }


def is_song_in_list(filtered_animes, songs, new_song):

    for existing_song in songs:
        print("comparing", existing_song["info"], new_song["info"])
        if existing_song["info"] == new_song["info"]:
            return True

    for anime in filtered_animes:
        for existing_song in anime["songs"]:
            if existing_song["info"] == new_song["info"]:
                return True

    return False


def filter_json_list(
    anime_search_filters, artist_search_filters, song_name_search_filters
):

    filtered_animes = []

    with open("expand.json", encoding="utf-8") as json_file:
        data = json.load(json_file)["questions"]
        for anime in data:
            anime_name = anime["name"]
            songs = []
            if and_filter:
                if anime_contains_anime_filters(anime, anime_search_filters):
                    for song in anime["songs"]:
                        if anime_contains_song_filters(
                            song,
                            artist_search_filters,
                            song_name_search_filters,
                        ):
                            song = format_song(song)
                            if filter_duplicate:
                                if not is_song_in_list(filtered_animes, songs, song):
                                    songs.append(song)
                            else:
                                songs.append(song)
            else:
                if anime_contains_anime_filters(anime, anime_search_filters):
                    for song in anime["songs"]:
                        song = format_song(song)
                        if filter_duplicate:
                            if not is_song_in_list(songs, song):
                                songs.append(song)
                        else:
                            songs.append(song)

                else:
                    for song in anime["songs"]:
                        if anime_contains_song_filters(
                            song,
                            artist_search_filters,
                            song_name_search_filters,
                        ):
                            song = format_song(song)
                            if filter_duplicate:
                                if not is_song_in_list(songs, song):
                                    songs.append(song)
                            else:
                                songs.append(song)

            if len(songs) > 0:
                filtered_animes.append({"name": anime_name, "songs": songs})

    return filtered_animes


def print_anime_list(filtered_animes):

    for anime in filtered_animes:
        print(anime["name"])
        for song in anime["songs"]:
            print(song["type"], ":", song["info"], "|", song["link"])
        print()


def create_workbook(filtered_animes):

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # First line value
    ws.cell(1, 1, "ID")
    ws.cell(1, 2, "Anime Name")
    ws.cell(1, 3, "Song Type")
    ws.cell(1, 4, "Song Info")
    ws.cell(1, 5, "mp3 Links")
    ws.cell(1, 6, "Full Versions")
    ws.cell(1, 7, "Rank")

    # Insert values
    row_iter = 2
    for anime in filtered_animes:
        for song in anime["songs"]:
            ws.cell(row_iter, 2, anime["name"])
            ws.cell(row_iter, 3, song["type"])
            ws.cell(row_iter, 4, song["info"]).hyperlink = song["link"]
            ws.cell(row_iter, 5, "Link").hyperlink = song["mp3_link"]
            ws.cell(row_iter, 6, "Link")
            row_iter += 1

    # Change width of column based on longest cell value
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max(
                    (dims.get(cell.column_letter, 0), len(str(cell.value)))
                )
    for col, value in dims.items():
        ws.column_dimensions[col].width = value + 7

    # general style
    gray_color = Color(rgb=cell_background_color)
    gray_background = PatternFill(patternType="solid", fgColor=gray_color)

    thin_border = Border(
        left=Side(style="thin", color=border_color),
        right=Side(style="thin", color=border_color),
        top=Side(style="thin", color=border_color),
        bottom=Side(style="thin", color=border_color),
    )

    for line in ws["A1:G" + str(row_iter - 1)]:
        for cell in line:
            cell.fill = gray_background
            cell.font = Font(size=rest_font_size, name=font_police)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    # Style for first line
    for line in ws["A1:G1"]:
        for cell in line:
            cell.font = Font(size=first_line_font_size, bold=True, name=font_police)

    # Blue color for links
    for line in ws["D2:E" + str(row_iter - 1)]:
        for cell in line:
            cell.font = Font(color=link_color)

    # Sorting property
    ws.auto_filter.ref = "A1:G" + str(row_iter - 1)

    wb.save(file_name)


if __name__ == "__main__":

    anime_search_filters = replace_regex_values(anime_search_filters)
    artist_search_filters = replace_regex_values(artist_search_filters)
    song_name_search_filters = replace_regex_values(song_name_search_filters)

    filtered_animes = filter_json_list(
        anime_search_filters, artist_search_filters, song_name_search_filters
    )

    print_anime_list(filtered_animes)

    create_workbook(filtered_animes)
