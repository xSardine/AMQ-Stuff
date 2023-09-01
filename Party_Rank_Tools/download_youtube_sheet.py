from pathlib import Path
import os, re
from openpyxl import load_workbook

# coordinates start at 0,0 top left
YOUTUBE_START_LINE = 1
YOUTUBE_LINKS_COLUMN = 4
SONG_NAME_COLUMN = 4
ARTIST_COLUMN = 3

DOWNLOAD_PATH = Path("downloaded/")
SHEET_NAME = "Sheet1"


def create_file_name_Windows(fileName, path, allowance=255):
    """
    Creates a windows-compliant filename by removing all bad characters
    and maintaining the windows path length limit (which by default is 255)
    """
    allowance -= (
        len(str(path)) + 1
    )  # by default, windows is sensitive to long total paths.
    bad_characters = re.compile(r"\\|/|<|>|:|\"|\||\?|\*|&|\^|\$|" + "\0")
    return create_file_name_common(fileName, path, bad_characters, allowance)


def create_file_name_common(fileName, path, bad_characters, allowance=255):
    if allowance > 255:
        allowance = 255  # on most common filesystems, including NTFS a filename can not exceed 255 characters
    # assign allowance for things that must be in the file name
    if allowance < 0:
        raise ValueError(
            """It is not possible to give a reasonable file name, due to length limitations.
        Consider changing location to somewhere with a shorter path."""
        )

    # make sure that user input doesn't contain bad characters
    fileName = bad_characters.sub("", fileName)
    print("\n\n", fileName, "\n")
    ret = ""
    for string in [fileName]:
        length = len(string)
        if allowance - length < 0:
            string = string[:allowance]
            length = len(string)
        ret += string
        allowance -= length

    print(path, ret)
    ret = path / Path(ret)

    return ret


def download_song(yt_link, filename):
    if "youtu" not in yt_link:
        return
    os.system(f"yt-dlp -ciw -S res,ext:mp4:m4a -S vcodec:avc1 {yt_link} -o {filename}")


if __name__ == "__main__":
    sheet_path = Path(".")
    sheet_list = list(sheet_path.glob("*.xlsx"))

    if len(sheet_list) > 0:
        for sheet in sheet_list:
            wb = load_workbook(sheet)
            ws = wb[SHEET_NAME]
            for i, row in enumerate(ws.iter_rows()):
                if i < YOUTUBE_START_LINE:
                    continue
                if row[YOUTUBE_LINKS_COLUMN].hyperlink:
                    yt_link = row[YOUTUBE_LINKS_COLUMN].hyperlink.target
                else:
                    yt_link = row[YOUTUBE_LINKS_COLUMN].value
                download_song(
                    yt_link,
                    create_file_name_Windows(
                        f"{row[SONG_NAME_COLUMN].value.replace(' ', '-')}_by_{row[ARTIST_COLUMN].value.replace(' ', '-')}",
                        DOWNLOAD_PATH,
                    ),
                )

