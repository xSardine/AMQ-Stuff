"""DL links in Husa's sheet
make sure to configure link_start_row and sheet_tab_name
@xSardine if problems"""

from pathlib import Path
import openpyxl
import re
import os

# Script Configuration
ignore_already_existing_name = True
output_path = Path("mp4")
# Script Configuration

# Sheet configuration
link_start_row = 2  # number of the row where I need to start getting links
link_position_column = 4  # position of the column containing webms in the sheet
file_name = "whatever, it's not necessary as long as it is in the same directory"
sheet_tab_name = "My_Sheet"  # name of the tab of the sheet
# Sheet configuration

# End of Configuration

# Search for any .xlsx file in the current directory or subdirectory
sheet_path = Path(".")
sheet_list = list(sheet_path.glob("**/*.xlsx"))

ffmpeg = "ffmpeg"

# Create webm directory if it doesn't exist
output_path.mkdir(parents=False, exist_ok=True)


def create_file_name_Windows(songTitle, path, allowance=255):
    """
    Creates a windows-compliant filename by removing all bad characters
    and maintaining the windows path length limit (which by default is 255)
    """
    allowance -= (
        len(str(path)) + 1
    )  # by default, windows is sensitive to long total paths.
    bad_characters = re.compile(r"\\|/|<|>|:|\"|\||\?|\*|&|\^|\$|" + "\0")
    return create_file_name_common(songTitle, path, bad_characters, allowance)


def create_file_name_common(songTitle, path, bad_characters, allowance=255):
    if allowance > 255:
        allowance = 255  # on most common filesystems, including NTFS a filename can not exceed 255 characters
    # assign allowance for things that must be in the file name
    allowance -= len("_-.mp4")  # accounting for separators (-_) for .webm
    if allowance < 0:
        raise ValueError(
            """It is not possible to give a reasonable file name, due to length limitations.
        Consider changing location to somewhere with a shorter path."""
        )
    # make sure that user input doesn't contain bad characters
    songTitle = bad_characters.sub("", songTitle)
    ret = ""
    for string in [songTitle]:
        length = len(string)
        if allowance - length < 0:
            string = string[:allowance]
            length = len(string)
        ret += string
        allowance -= length
        if allowance - 1 > 1:
            ret += "-"
        else:
            break
    else:
        ret = ret[:-1]  # removes last "-"
    ret = path.joinpath(Path(ret + ".mp4"))

    return str(ret)


def execute_command(command):
    os.system(command)


if len(sheet_list) > 0:  # If I found an .xlsx

    # Open the first one
    wb = openpyxl.load_workbook(sheet_list[0])

    ws = wb[sheet_tab_name]  # access the right tab
    flag_exception = True
    if ignore_already_existing_name:
        ignore_parameter = "-y"
    else:
        ignored_parameter = "-n"

    while flag_exception:  # loop on every link until I end up on a non-link cell
        song_name = ""
        try:
            song_name = ws.cell(
                row=link_start_row, column=link_position_column
            ).value.replace(" ", "_")
            link = ws.cell(
                row=link_start_row, column=link_position_column
            ).hyperlink.target
            command = [
                "%s" % ffmpeg,
                ignore_parameter,
                "-i",
                link,
                "-c:a",
                "aac",
                "-c:v",
                "libx264",
                "-map_metadata",
                "-1",
                "-map_chapters",
                "-1 " '"%s"' % create_file_name_Windows(song_name, output_path),
            ]
            print(song_name, "->", link)
            print(" ".join(command))
            execute_command(" ".join(command))
            link_start_row += 1
        except:
            flag_exception = False
            if len(song_name) > 0:
                print("Failed for", link, "(" + song_name + ")")
    print("Script is Done :)")
